"""
SUITE DE PRUEBAS - Dashboard Brigadas MSC
==========================================
Ejecutar con: python test_suite.py
Requiere: pip install openpyxl

Este script calcula los valores esperados directamente desde el Excel
para comparar contra lo que muestra el dashboard en el navegador.
"""

import sys
import openpyxl
from collections import defaultdict, Counter
from datetime import datetime, date

# ── CONFIGURACIÓN ─────────────────────────────────────────────────────────────
EXCEL_PATH = input("Ruta al archivo Excel (Enter para usar el default): ").strip()
if not EXCEL_PATH:
    EXCEL_PATH = "Capacitación_URE_BV-GPR-LE_.xlsx"

print(f"\nCargando {EXCEL_PATH}...")
try:
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
except FileNotFoundError:
    print(f"ERROR: No se encontró el archivo '{EXCEL_PATH}'")
    sys.exit(1)

ws_s = wb['BD_seg']
ws_p = wb['BD_personal']

# ── PARSEO COMPLETO ───────────────────────────────────────────────────────────
print("Parseando datos...\n")

personal = {}
for r in range(2, ws_p.max_row + 1):
    ci = ws_p.cell(row=r, column=2).value
    if not ci: continue
    personal[str(ci).strip()] = {
        'nombre':     str(ws_p.cell(row=r, column=3).value or '').strip(),
        'estado':     str(ws_p.cell(row=r, column=4).value or '').strip(),
        'tipo_bv':    str(ws_p.cell(row=r, column=7).value or '').strip(),
        'nivel':      str(ws_p.cell(row=r, column=9).value or '').strip(),
        'area':       str(ws_p.cell(row=r, column=10).value or '').strip(),
        'grupo':      str(ws_p.cell(row=r, column=12).value or '').strip(),
        'supervisor': str(ws_p.cell(row=r, column=18).value or '').strip(),
    }

raw = []
mod_lec = defaultdict(set)
for r in range(2, ws_s.max_row + 1):
    nro = ws_s.cell(row=r, column=1).value
    if nro is None: continue
    ci     = str(ws_s.cell(row=r, column=2).value or '').strip()
    modulo = str(ws_s.cell(row=r, column=5).value or '').strip()
    cod    = str(ws_s.cell(row=r, column=17).value or '').strip()
    nota_r = ws_s.cell(row=r, column=10).value
    hrs_r  = ws_s.cell(row=r, column=12).value
    fecha  = ws_s.cell(row=r, column=13).value
    tipo   = str(ws_s.cell(row=r, column=18).value or '').strip()
    area   = str(ws_s.cell(row=r, column=19).value or '').strip()

    if modulo and cod: mod_lec[modulo].add(cod)

    hrs = 0.0
    try: hrs = float(str(hrs_r).replace(',', '.')) if hrs_r else 0.0
    except: pass

    nota = None
    try: nota = float(nota_r) if nota_r is not None and str(nota_r).strip() else None
    except: pass

    fecha_dt = None
    if isinstance(fecha, (datetime, date)):
        fecha_dt = fecha if isinstance(fecha, datetime) else datetime.combine(fecha, datetime.min.time())
    elif fecha:
        try: fecha_dt = datetime.strptime(str(fecha).strip()[:10], '%Y-%m-%d')
        except: pass

    raw.append({
        'ci': ci, 'modulo': modulo, 'cod': cod,
        'nota': nota, 'hrs': hrs, 'fecha': fecha_dt,
        'tipo': tipo, 'area': area, 'asistio': hrs > 0,
        'row': r
    })

mod_total = {m: len(s) for m, s in mod_lec.items()}

PASS = "  ✓ PASS"
FAIL = "  ✗ FAIL"
SEP  = "─" * 60
results = {'pass': 0, 'fail': 0}

def check(label, got, expected, tol=0.05):
    ok = abs(float(got) - float(expected)) <= tol
    symbol = PASS if ok else FAIL
    results['pass' if ok else 'fail'] += 1
    print(f"{symbol}  {label}")
    if not ok:
        print(f"        Obtenido: {got}  |  Esperado: {expected}")

def check_str(label, got, expected):
    ok = str(got).strip() == str(expected).strip()
    symbol = PASS if ok else FAIL
    results['pass' if ok else 'fail'] += 1
    print(f"{symbol}  {label}")
    if not ok:
        print(f"        Obtenido: '{got}'  |  Esperado: '{expected}'")

def section(title):
    print(f"\n{SEP}")
    print(f"  {title}")
    print(SEP)

# ══════════════════════════════════════════════════════════════════════════════
# PRUEBA 1: INTEGRIDAD TOTAL DE DATOS
# ══════════════════════════════════════════════════════════════════════════════
section("PRUEBA 1 — Integridad total de datos (tab Verificar datos)")
print("  Comparar estos números con la sección 'Verificar datos' del dashboard\n")

total_rows    = len(raw)
rows_asis     = sum(1 for d in raw if d['asistio'])
rows_falta    = sum(1 for d in raw if not d['asistio'])
cis_unicos    = len(set(d['ci'] for d in raw if d['ci']))
total_hrs     = sum(d['hrs'] for d in raw if d['asistio'])
total_personal= len(personal)
total_modulos = len(mod_total)

print(f"  Total registros BD_seg:        {total_rows}")
print(f"  Registros con TIEMPO > 0:      {rows_asis}")
print(f"  Registros sin tiempo (falta):  {rows_falta}")
print(f"  CIs únicos en BD_seg:          {cis_unicos}")
print(f"  Horas totales programa:        {total_hrs:.1f}")
print(f"  Brigadistas en BD_personal:    {total_personal}")
print(f"  Módulos únicos en programa:    {total_modulos}")

# ══════════════════════════════════════════════════════════════════════════════
# PRUEBA 2: FILTRO TEMPORAL
# ══════════════════════════════════════════════════════════════════════════════
section("PRUEBA 2 — Filtro temporal por año")
print("  En el dashboard: seleccioná cada año en la barra de período\n")

for year in [2024, 2025, 2026]:
    yr = [d for d in raw if d['fecha'] and d['fecha'].year == year]
    brigadas = len(set(d['ci'] for d in yr if d['ci']))
    asis = sum(1 for d in yr if d['asistio'])
    hrs  = sum(d['hrs'] for d in yr if d['asistio'])
    pct_asis = round(asis / len(yr) * 100) if yr else 0
    print(f"  {year}:")
    print(f"    Registros:          {len(yr)}")
    print(f"    Brigadistas únicos: {brigadas}")
    print(f"    Asistencias:        {asis} / {len(yr)} = {pct_asis}%")
    print(f"    Horas totales:      {hrs:.1f}")

# ══════════════════════════════════════════════════════════════════════════════
# PRUEBA 3: BRIGADISTA INDIVIDUAL — 3 PERSONAS DISTINTAS
# ══════════════════════════════════════════════════════════════════════════════
section("PRUEBA 3 — Consulta individual (tab Brigadista)")
print("  Buscá cada CI en el dashboard y compará los valores\n")

# Pick 3 CIs: top-hrs, mid, someone with absences
test_cis = ['4151385', '8643511']
# Find one person with at least 1 absence
for d in raw:
    if not d['asistio'] and d['ci'] not in test_cis and d['ci']:
        test_cis.append(d['ci'])
        break
test_cis = list(dict.fromkeys(test_cis))[:3]

for ci in test_cis:
    p_rows = [d for d in raw if d['ci'] == ci]
    if not p_rows: continue
    prof = personal.get(ci, {})
    ses   = len(p_rows)
    asis  = sum(1 for d in p_rows if d['asistio'])
    hrs   = sum(d['hrs'] for d in p_rows if d['asistio'])
    notas = [d['nota'] for d in p_rows if d['nota'] is not None]
    pct_a = round(asis / ses * 100) if ses else 0
    prom  = round(sum(notas) / len(notas), 1) if notas else None
    mods  = set(d['modulo'] for d in p_rows if d['modulo'])

    print(f"  CI: {ci} — {prof.get('nombre','SIN PERFIL')}")
    print(f"    Área:         {prof.get('area','—')}")
    print(f"    Grupo:        {prof.get('grupo','—')}")
    print(f"    Supervisor:   {prof.get('supervisor','—')}")
    print(f"    Tipo BV:      {prof.get('tipo_bv','—')}")
    print(f"    Sesiones:     {ses}")
    print(f"    Asistencias:  {asis} ({pct_a}%)")
    print(f"    Horas totales:{hrs:.1f}")
    print(f"    Nota promedio:{prom if prom else '—'}")
    print(f"    Módulos:      {sorted(mods)}")
    # Per module
    print(f"    Detalle por módulo:")
    by_mod = defaultdict(lambda: {'lec_ok':0,'hrs':0,'notas':[]})
    for d in p_rows:
        if d['modulo']:
            if d['asistio']: by_mod[d['modulo']]['lec_ok'] += 1; by_mod[d['modulo']]['hrs'] += d['hrs']
            if d['nota'] is not None: by_mod[d['modulo']]['notas'].append(d['nota'])
    for m, v in sorted(by_mod.items()):
        lt = mod_total.get(m, 1)
        pct_mod = round(v['lec_ok'] / lt * 100)
        na = round(sum(v['notas'])/len(v['notas']),1) if v['notas'] else None
        print(f"      {m}: {v['lec_ok']}/{lt} lec ({pct_mod}%) | {v['hrs']:.1f}hr | nota:{na or '—'}")
    print()

# ══════════════════════════════════════════════════════════════════════════════
# PRUEBA 4: VISTA POR GRUPO
# ══════════════════════════════════════════════════════════════════════════════
section("PRUEBA 4 — Vista por grupo")
print("  En el dashboard: tab Grupos, seleccioná Área Mina y Grupo A\n")

area_test = 'Mina'; grupo_test = 'A'
cis_grupo = [ci for ci, p in personal.items() if p['area'] == area_test and p['grupo'] == grupo_test]
g_rows = [d for d in raw if d['ci'] in cis_grupo]
g_ses  = len(g_rows)
g_asis = sum(1 for d in g_rows if d['asistio'])
g_hrs  = sum(d['hrs'] for d in g_rows if d['asistio'])
g_notas= [d['nota'] for d in g_rows if d['nota'] is not None]
g_pct  = round(g_asis / g_ses * 100) if g_ses else 0
g_nota = round(sum(g_notas)/len(g_notas),1) if g_notas else None

by_person_asis = {}
for ci in cis_grupo:
    pr = [d for d in g_rows if d['ci'] == ci]
    if pr: by_person_asis[ci] = round(sum(1 for d in pr if d['asistio']) / len(pr) * 100)

best_ci = max(by_person_asis, key=by_person_asis.get) if by_person_asis else None
worst_ci= min(by_person_asis, key=by_person_asis.get) if by_person_asis else None

print(f"  Área: {area_test} | Grupo: {grupo_test}")
print(f"  Brigadistas en grupo:    {len(cis_grupo)}")
print(f"  % Asistencia del grupo:  {g_pct}%")
print(f"  Nota promedio:           {g_nota or '—'}")
print(f"  Horas totales:           {g_hrs:.1f}")
print(f"  Mayor asistencia: {personal.get(best_ci,{}).get('nombre','—')} ({by_person_asis.get(best_ci,'—')}%)")
print(f"  Menor asistencia: {personal.get(worst_ci,{}).get('nombre','—')} ({by_person_asis.get(worst_ci,'—')}%)")

# ══════════════════════════════════════════════════════════════════════════════
# PRUEBA 5: PROGRAMA POR ÁREA
# ══════════════════════════════════════════════════════════════════════════════
section("PRUEBA 5 — Programa por área (tab Por Área)")
print("  Comparar con la sección 'Por Área' del dashboard\n")

for area in ['Mina', 'Planta', 'Servicios']:
    cis_area = [ci for ci, p in personal.items() if p['area'] == area]
    cis_bvm  = [ci for ci in cis_area if personal[ci].get('tipo_bv') == 'BV-M']
    a_rows   = [d for d in raw if d['ci'] in cis_area]
    a_ses    = len(a_rows); a_asis = sum(1 for d in a_rows if d['asistio'])
    pct_a    = round(a_asis / a_ses * 100) if a_ses else 0
    grupos   = set(personal[ci]['grupo'] for ci in cis_area if personal[ci].get('grupo'))
    print(f"  Área {area}:")
    print(f"    Total brigadistas:  {len(cis_area)}")
    print(f"    BV-M:               {len(cis_bvm)}")
    print(f"    Grupos:             {len(grupos)}")
    print(f"    % Asistencia área:  {pct_a}%")

# ══════════════════════════════════════════════════════════════════════════════
# PRUEBA 6: NIVEL INICIAL
# ══════════════════════════════════════════════════════════════════════════════
section("PRUEBA 6 — Nivel Inicial (tab Nv. Inicial)")

cis_ini = [ci for ci, p in personal.items() if p.get('nivel') == 'Inicial' or p.get('tipo_bv') == 'BV-B']
ini_rows= [d for d in raw if d['ci'] in cis_ini]
ini_ses = len(ini_rows); ini_asis = sum(1 for d in ini_rows if d['asistio'])
ini_hrs = sum(d['hrs'] for d in ini_rows if d['asistio'])
ini_notas=[d['nota'] for d in ini_rows if d['nota'] is not None]
ini_pct = round(ini_asis / ini_ses * 100) if ini_ses else 0
ini_nota= round(sum(ini_notas)/len(ini_notas),1) if ini_notas else None

by_ini_asis = {}
for ci in cis_ini:
    pr = [d for d in ini_rows if d['ci'] == ci]
    if pr: by_ini_asis[ci] = round(sum(1 for d in pr if d['asistio']) / len(pr) * 100)
best_ini  = max(by_ini_asis, key=by_ini_asis.get) if by_ini_asis else None
worst_ini = min(by_ini_asis, key=by_ini_asis.get) if by_ini_asis else None

print(f"  Total brigadistas nivel inicial: {len(cis_ini)}")
print(f"  % Asistencia:                    {ini_pct}%")
print(f"  Nota promedio:                   {ini_nota or '—'}")
print(f"  Horas totales:                   {ini_hrs:.1f}")
print(f"  Mayor asistencia: {personal.get(best_ini,{}).get('nombre','—')} ({by_ini_asis.get(best_ini,'—')}%)")
print(f"  Menor asistencia: {personal.get(worst_ini,{}).get('nombre','—')} ({by_ini_asis.get(worst_ini,'—')}%)")

# ══════════════════════════════════════════════════════════════════════════════
# PRUEBA 7: TENDENCIA MENSUAL
# ══════════════════════════════════════════════════════════════════════════════
section("PRUEBA 7 — Tendencia temporal (tab Tendencia)")
print("  Comparar los primeros y últimos meses con los gráficos de tendencia\n")

by_month = defaultdict(lambda: {'asis':0,'total':0,'hrs':0})
for d in raw:
    if not d['fecha']: continue
    k = f"{d['fecha'].year}-{d['fecha'].month:02d}"
    by_month[k]['total'] += 1
    if d['asistio']: by_month[k]['asis'] += 1; by_month[k]['hrs'] += d['hrs']

print(f"  {'Mes':<12} {'Registros':>10} {'Asistencia':>12} {'Horas':>10}")
print(f"  {'─'*12} {'─'*10} {'─'*12} {'─'*10}")
for k in sorted(by_month.keys()):
    v = by_month[k]
    pct_m = round(v['asis'] / v['total'] * 100) if v['total'] else 0
    print(f"  {k:<12} {v['total']:>10} {str(pct_m)+'%':>12} {v['hrs']:>10.1f}")

# ══════════════════════════════════════════════════════════════════════════════
# PRUEBA 8: EXCEL MANUAL — FÓRMULAS PARA VERIFICAR
# ══════════════════════════════════════════════════════════════════════════════
section("PRUEBA 8 — Fórmulas Excel para verificación independiente")
print("  Abrí BD_seg en Excel y ejecutá estas fórmulas para verificar\n")

print("  [1] Total de filas con datos (debe dar 1916):")
print("      =COUNTA(A2:A10000)")
print()
print("  [2] Total filas con TIEMPO > 0 (debe dar 1886):")
print("      =COUNTIF(L2:L10000,\">0\")")
print()
print("  [3] Horas totales programa (debe dar 3863.5):")
print("      =SUMIF(L2:L10000,\">0\",L2:L10000)")
print()
print("  [4] Registros de un CI específico — ej. 4151385 (debe dar 42):")
print("      =COUNTIF(B2:B10000,\"4151385\")")
print()
print("  [5] Horas de CI 4151385 (debe dar 91.0):")
print("      =SUMIF(B2:B10000,\"4151385\",L2:L10000)")
print()
print("  [6] Registros del año 2024 (debe dar 720):")
print("      =COUNTIFS(M2:M10000,\">=\"&DATE(2024,1,1),M2:M10000,\"<=\"&DATE(2024,12,31))")
print()
print("  [7] Registros del año 2025 (debe dar 797):")
print("      =COUNTIFS(M2:M10000,\">=\"&DATE(2025,1,1),M2:M10000,\"<=\"&DATE(2025,12,31))")

# ══════════════════════════════════════════════════════════════════════════════
# PRUEBA 9: CASOS LÍMITE
# ══════════════════════════════════════════════════════════════════════════════
section("PRUEBA 9 — Casos límite y situaciones inesperadas")

# CIs en BD_seg sin perfil en BD_personal
sin_perfil = set(d['ci'] for d in raw if d['ci'] and d['ci'] not in personal)
print(f"  CIs en BD_seg sin perfil en BD_personal: {len(sin_perfil)}")
if sin_perfil: print(f"    → {sin_perfil}")

# CIs en BD_personal sin actividad
sin_actividad = set(personal.keys()) - set(d['ci'] for d in raw if d['ci'])
print(f"  Brigadistas en BD_personal sin ninguna actividad: {len(sin_actividad)}")
print(f"    (Estos aparecen como perfiles vacíos si se buscan en el dashboard)")

# Notas fuera de rango 0-100
notas_raras = [(d['ci'], d['nota'], d['row']) for d in raw if d['nota'] is not None and (d['nota'] < 0 or d['nota'] > 100)]
print(f"  Notas fuera del rango 0-100: {len(notas_raras)}")
if notas_raras: print(f"    → {notas_raras[:5]}")

# Horas negativas o extremas
hrs_raras = [(d['ci'], d['hrs'], d['row']) for d in raw if d['hrs'] > 24]
print(f"  Registros con TIEMPO > 24 hrs (sospechoso): {len(hrs_raras)}")
if hrs_raras: print(f"    → {hrs_raras[:5]}")

# Registros sin módulo
sin_modulo = sum(1 for d in raw if not d['modulo'])
print(f"  Registros sin módulo asignado: {sin_modulo}")

# Fechas futuras (más allá de hoy)
hoy = datetime.now()
fechas_futuras = [d for d in raw if d['fecha'] and d['fecha'] > hoy]
print(f"  Registros con fecha futura: {len(fechas_futuras)}")
if fechas_futuras: print(f"    → {[(d['ci'], d['fecha'].date()) for d in fechas_futuras[:3]]}")

# ══════════════════════════════════════════════════════════════════════════════
# RESUMEN FINAL
# ══════════════════════════════════════════════════════════════════════════════
section("RESUMEN")
total_checks = results['pass'] + results['fail']
print(f"  Checks automáticos: {results['pass']}/{total_checks} pasaron")
print()
print("  PASOS MANUALES PENDIENTES (verificar visualmente en el dashboard):")
print("  □ Prueba 1: Abrir tab 'Verificar datos', todos deben mostrar ✓")
print("  □ Prueba 2: Cambiar período a 2024/2025/2026, comparar totales con la tabla de arriba")
print("  □ Prueba 3: Buscar CI 4151385, verificar horas=91.0, asistencia=100%, nota≈97.5")
print("  □ Prueba 3: Buscar CI 8643511, comparar todos los campos con la tabla de arriba")
print("  □ Prueba 4: Grupos → Mina → A, comparar % asistencia y mejor/peor")
print("  □ Prueba 5: Por Área, comparar totales por área")
print("  □ Prueba 6: Nv. Inicial, comparar total y asistencia")
print("  □ Prueba 7: Tendencia, verificar que el primer mes sea 2024-01 y el último 2026-05")
print("  □ Prueba 8: Ejecutar fórmulas en Excel y comparar con valores esperados")
print()
if results['fail'] == 0:
    print("  ✓ Todos los checks automáticos pasaron.")
else:
    print(f"  ✗ {results['fail']} check(s) fallaron — revisá los detalles arriba.")

print(f"\n{'─'*60}")
print("  Script finalizado. Guardá este output para comparar con el dashboard.")
print(f"{'─'*60}\n")
